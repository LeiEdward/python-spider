# -*- coding: utf-8 -*-

# 參數設定
VERIFYPASS = False
SHOWCHROME = True
VERIFYCOUNT = 1
VERSION = 'v4.2'

# 舊版
SEARCH_A = ['租屋']
WEBTYPE_A = ['租屋','店面出租', '辦公出租','住辦出租','廠房出租','土地出租']
# 新版
SEARCH_B = ['店面出租', '店面出售','辦公出租','住辦出租','辦公出售','住辦出售','廠房出租','土地出租','廠房出售','土地出售']
WEBTYPE_B = ['中古屋', '店面出售','辦公出售','住辦出售','廠房出售']

SEARCH_C = ['中古屋']
WEBTYPE_C = ['中古屋']

WEBTYPE_D = ['土地出售']

# 時間
from datetime import datetime,timezone,timedelta
import time
import ntplib


# 基本引入
import json
import shutil
import os
import io
import requests
import sys
from getmac import get_mac_address
import hashlib
import inspect
import logging
import random
import webbrowser

# 使用內建的 urllib.request 裡的 urlopen 這個功能來送出網址
from urllib.request import urlopen, urlretrieve
from urllib.error import HTTPError
from bs4 import BeautifulSoup
#如果是 MAC 電腦, 請務必加入下面兩行, 因為 MAC 有視 https 的 ssl 證書無效的 bug
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

import pandas as pd

import warnings
#忽略掉對ignore的warning
warnings.filterwarnings('ignore')

import styleframe
import glob
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing import image

# chrome模擬器
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ['enable-automation', 'enable-logging'])

# fake_useragent
from fake_useragent import UserAgent

# 視窗設計
import tkinter as tk
import tkinter.ttk as tt

# 圖片轉文字
import pytesseract
from PIL import Image
pytesseract.pytesseract.tesseract_cmd = './Tesseract-OCR/tesseract.exe'


# google firestore
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore



# 初始化firebase，注意不能重複初始化
cred = credentials.Certificate("./serviceAccountKey.json")
firebase_admin.initialize_app(cred)


# 圖片辨識測試
# img_path = "phone_img_"+ str(datetime.date.today())
# dirFiles = os.listdir(img_path)
# iii = 0
# for cell in list(dirFiles):
#     im = Image.open(img_path + "/" + dirFiles[iii])
#     (x,y) = im.size #read image size
#     x_s = 150 #define standard width
#     y_s = y * x_s / x #calc height based on standard width
#     out = im.resize((x_s, int (y_s)),Image.ANTIALIAS) #resize image with high-quality
#     out.save(img_path + "/" + dirFiles[iii])
#     phone_t = pytesseract.image_to_string(out)
#     print(x_s, img_path + "/" + dirFiles[iii], phone_t)
#     iii = iii + 1

import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

Taiwan = json.loads('[{"districts":[{"zip":"100","name":"中正區"},{"zip":"103","name":"大同區"},{"zip":"104","name":"中山區"},{"zip":"105","name":"松山區"},{"zip":"106","name":"大安區"},{"zip":"108","name":"萬華區"},{"zip":"110","name":"信義區"},{"zip":"111","name":"士林區"},{"zip":"112","name":"北投區"},{"zip":"114","name":"內湖區"},{"zip":"115","name":"南港區"},{"zip":"116","name":"文山區"}],"name":"台北市"},{"districts":[{"zip":"200","name":"仁愛區"},{"zip":"201","name":"信義區"},{"zip":"202","name":"中正區"},{"zip":"203","name":"中山區"},{"zip":"204","name":"安樂區"},{"zip":"205","name":"暖暖區"},{"zip":"206","name":"七堵區"}],"name":"基隆市"},{"districts":[{"zip":"207","name":"萬里區"},{"zip":"208","name":"金山區"},{"zip":"220","name":"板橋區"},{"zip":"221","name":"汐止區"},{"zip":"222","name":"深坑區"},{"zip":"223","name":"石碇區"},{"zip":"224","name":"瑞芳區"},{"zip":"226","name":"平溪區"},{"zip":"227","name":"雙溪區"},{"zip":"228","name":"貢寮區"},{"zip":"231","name":"新店區"},{"zip":"232","name":"坪林區"},{"zip":"233","name":"烏來區"},{"zip":"234","name":"永和區"},{"zip":"235","name":"中和區"},{"zip":"236","name":"土城區"},{"zip":"237","name":"三峽區"},{"zip":"238","name":"樹林區"},{"zip":"239","name":"鶯歌區"},{"zip":"241","name":"三重區"},{"zip":"242","name":"新莊區"},{"zip":"243","name":"泰山區"},{"zip":"244","name":"林口區"},{"zip":"247","name":"蘆洲區"},{"zip":"248","name":"五股區"},{"zip":"249","name":"八里區"},{"zip":"251","name":"淡水區"},{"zip":"252","name":"三芝區"},{"zip":"253","name":"石門區"}],"name":"新北市"},{"districts":[{"zip":"209","name":"南竿鄉"},{"zip":"210","name":"北竿鄉"},{"zip":"211","name":"莒光鄉"},{"zip":"212","name":"東引鄉"}],"name":"連江縣"},{"districts":[{"zip":"260","name":"宜蘭市"},{"zip":"263","name":"壯圍鄉"},{"zip":"261","name":"頭城鎮"},{"zip":"262","name":"礁溪鄉"},{"zip":"264","name":"員山鄉"},{"zip":"265","name":"羅東鎮"},{"zip":"266","name":"三星鄉"},{"zip":"267","name":"大同鄉"},{"zip":"268","name":"五結鄉"},{"zip":"269","name":"冬山鄉"},{"zip":"270","name":"蘇澳鎮"},{"zip":"272","name":"南澳鄉"},{"zip":"290","name":"釣魚台"}],"name":"宜蘭縣"},{"districts":[{"zip":"290","name":"釣魚台"}],"name":"釣魚台"},{"districts":[{"zip":"300","name":"東區"},{"zip":"300","name":"北區"},{"zip":"300","name":"香山區"}],"name":"新竹市"},{"districts":[{"zip":"308","name":"寶山鄉"},{"zip":"302","name":"竹北市"},{"zip":"303","name":"湖口鄉"},{"zip":"304","name":"新豐鄉"},{"zip":"305","name":"新埔鎮"},{"zip":"306","name":"關西鎮"},{"zip":"307","name":"芎林鄉"},{"zip":"310","name":"竹東鎮"},{"zip":"311","name":"五峰鄉"},{"zip":"312","name":"橫山鄉"},{"zip":"313","name":"尖石鄉"},{"zip":"314","name":"北埔鄉"},{"zip":"315","name":"峨眉鄉"}],"name":"新竹縣"},{"districts":[{"zip":"320","name":"中壢區"},{"zip":"324","name":"平鎮區"},{"zip":"325","name":"龍潭區"},{"zip":"326","name":"楊梅區"},{"zip":"327","name":"新屋區"},{"zip":"328","name":"觀音區"},{"zip":"330","name":"桃園區"},{"zip":"333","name":"龜山區"},{"zip":"334","name":"八德區"},{"zip":"335","name":"大溪區"},{"zip":"336","name":"復興區"},{"zip":"337","name":"大園區"},{"zip":"338","name":"蘆竹區"}],"name":"桃園市"},{"districts":[{"zip":"350","name":"竹南鎮"},{"zip":"351","name":"頭份市"},{"zip":"352","name":"三灣鄉"},{"zip":"353","name":"南庄鄉"},{"zip":"354","name":"獅潭鄉"},{"zip":"356","name":"後龍鎮"},{"zip":"357","name":"通霄鎮"},{"zip":"358","name":"苑裡鎮"},{"zip":"360","name":"苗栗市"},{"zip":"361","name":"造橋鄉"},{"zip":"362","name":"頭屋鄉"},{"zip":"363","name":"公館鄉"},{"zip":"364","name":"大湖鄉"},{"zip":"365","name":"泰安鄉"},{"zip":"366","name":"銅鑼鄉"},{"zip":"367","name":"三義鄉"},{"zip":"368","name":"西湖鄉"},{"zip":"369","name":"卓蘭鎮"}],"name":"苗栗縣"},{"districts":[{"zip":"400","name":"中區"},{"zip":"401","name":"東區"},{"zip":"402","name":"南區"},{"zip":"403","name":"西區"},{"zip":"404","name":"北區"},{"zip":"406","name":"北屯區"},{"zip":"407","name":"西屯區"},{"zip":"408","name":"南屯區"},{"zip":"411","name":"太平區"},{"zip":"412","name":"大里區"},{"zip":"413","name":"霧峰區"},{"zip":"414","name":"烏日區"},{"zip":"420","name":"豐原區"},{"zip":"421","name":"后里區"},{"zip":"422","name":"石岡區"},{"zip":"423","name":"東勢區"},{"zip":"424","name":"和平區"},{"zip":"426","name":"新社區"},{"zip":"427","name":"潭子區"},{"zip":"428","name":"大雅區"},{"zip":"429","name":"神岡區"},{"zip":"432","name":"大肚區"},{"zip":"433","name":"沙鹿區"},{"zip":"434","name":"龍井區"},{"zip":"435","name":"梧棲區"},{"zip":"436","name":"清水區"},{"zip":"437","name":"大甲區"},{"zip":"438","name":"外埔區"},{"zip":"439","name":"大安區"}],"name":"台中市"},{"districts":[{"zip":"500","name":"彰化市"},{"zip":"502","name":"芬園鄉"},{"zip":"503","name":"花壇鄉"},{"zip":"504","name":"秀水鄉"},{"zip":"505","name":"鹿港鎮"},{"zip":"506","name":"福興鄉"},{"zip":"507","name":"線西鄉"},{"zip":"508","name":"和美鎮"},{"zip":"509","name":"伸港鄉"},{"zip":"510","name":"員林市"},{"zip":"511","name":"社頭鄉"},{"zip":"512","name":"永靖鄉"},{"zip":"513","name":"埔心鄉"},{"zip":"514","name":"溪湖鎮"},{"zip":"515","name":"大村鄉"},{"zip":"516","name":"埔鹽鄉"},{"zip":"520","name":"田中鎮"},{"zip":"521","name":"北斗鎮"},{"zip":"522","name":"田尾鄉"},{"zip":"523","name":"埤頭鄉"},{"zip":"524","name":"溪州鄉"},{"zip":"525","name":"竹塘鄉"},{"zip":"526","name":"二林鎮"},{"zip":"527","name":"大城鄉"},{"zip":"528","name":"芳苑鄉"},{"zip":"530","name":"二水鄉"}],"name":"彰化縣"},{"districts":[{"zip":"540","name":"南投市"},{"zip":"541","name":"中寮鄉"},{"zip":"542","name":"草屯鎮"},{"zip":"544","name":"國姓鄉"},{"zip":"545","name":"埔里鎮"},{"zip":"546","name":"仁愛鄉"},{"zip":"551","name":"名間鄉"},{"zip":"552","name":"集集鎮"},{"zip":"553","name":"水里鄉"},{"zip":"555","name":"魚池鄉"},{"zip":"556","name":"信義鄉"},{"zip":"557","name":"竹山鎮"},{"zip":"558","name":"鹿谷鄉"}],"name":"南投縣"},{"districts":[{"zip":"600","name":"西區"},{"zip":"600","name":"東區"}],"name":"嘉義市"},{"districts":[{"zip":"602","name":"番路鄉"},{"zip":"603","name":"梅山鄉"},{"zip":"604","name":"竹崎鄉"},{"zip":"605","name":"阿里山鄉"},{"zip":"606","name":"中埔鄉"},{"zip":"607","name":"大埔鄉"},{"zip":"608","name":"水上鄉"},{"zip":"611","name":"鹿草鄉"},{"zip":"612","name":"太保市"},{"zip":"613","name":"朴子市"},{"zip":"614","name":"東石鄉"},{"zip":"615","name":"六腳鄉"},{"zip":"616","name":"新港鄉"},{"zip":"621","name":"民雄鄉"},{"zip":"622","name":"大林鎮"},{"zip":"623","name":"溪口鄉"},{"zip":"624","name":"義竹鄉"},{"zip":"625","name":"布袋鎮"}],"name":"嘉義縣"},{"districts":[{"zip":"630","name":"斗南鎮"},{"zip":"631","name":"大埤鄉"},{"zip":"632","name":"虎尾鎮"},{"zip":"633","name":"土庫鎮"},{"zip":"634","name":"褒忠鄉"},{"zip":"635","name":"東勢鄉"},{"zip":"636","name":"台西鄉"},{"zip":"637","name":"崙背鄉"},{"zip":"638","name":"麥寮鄉"},{"zip":"640","name":"斗六市"},{"zip":"643","name":"林內鄉"},{"zip":"646","name":"古坑鄉"},{"zip":"647","name":"莿桐鄉"},{"zip":"648","name":"西螺鎮"},{"zip":"649","name":"二崙鄉"},{"zip":"651","name":"北港鎮"},{"zip":"652","name":"水林鄉"},{"zip":"653","name":"口湖鄉"},{"zip":"654","name":"四湖鄉"},{"zip":"655","name":"元長鄉"}],"name":"雲林縣"},{"districts":[{"zip":"700","name":"中西區"},{"zip":"701","name":"東區"},{"zip":"702","name":"南區"},{"zip":"704","name":"北區"},{"zip":"708","name":"安平區"},{"zip":"709","name":"安南區"},{"zip":"710","name":"永康區"},{"zip":"711","name":"歸仁區"},{"zip":"712","name":"新化區"},{"zip":"713","name":"左鎮區"},{"zip":"714","name":"玉井區"},{"zip":"715","name":"楠西區"},{"zip":"716","name":"南化區"},{"zip":"717","name":"仁德區"},{"zip":"718","name":"關廟區"},{"zip":"719","name":"龍崎區"},{"zip":"720","name":"官田區"},{"zip":"721","name":"麻豆區"},{"zip":"722","name":"佳里區"},{"zip":"723","name":"西港區"},{"zip":"724","name":"七股區"},{"zip":"725","name":"將軍區"},{"zip":"726","name":"學甲區"},{"zip":"727","name":"北門區"},{"zip":"730","name":"新營區"},{"zip":"731","name":"後壁區"},{"zip":"732","name":"白河區"},{"zip":"733","name":"東山區"},{"zip":"734","name":"六甲區"},{"zip":"735","name":"下營區"},{"zip":"736","name":"柳營區"},{"zip":"737","name":"鹽水區"},{"zip":"741","name":"善化區"},{"zip":"744","name":"新市區"},{"zip":"742","name":"大內區"},{"zip":"743","name":"山上區"},{"zip":"745","name":"安定區"}],"name":"台南市"},{"districts":[{"zip":"800","name":"新興區"},{"zip":"801","name":"前金區"},{"zip":"802","name":"苓雅區"},{"zip":"803","name":"鹽埕區"},{"zip":"804","name":"鼓山區"},{"zip":"805","name":"旗津區"},{"zip":"806","name":"前鎮區"},{"zip":"807","name":"三民區"},{"zip":"811","name":"楠梓區"},{"zip":"812","name":"小港區"},{"zip":"813","name":"左營區"},{"zip":"814","name":"仁武區"},{"zip":"815","name":"大社區"},{"zip":"817","name":"東沙群島"},{"zip":"819","name":"南沙群島"},{"zip":"820","name":"岡山區"},{"zip":"821","name":"路竹區"},{"zip":"822","name":"阿蓮區"},{"zip":"823","name":"田寮區"},{"zip":"824","name":"燕巢區"},{"zip":"825","name":"橋頭區"},{"zip":"826","name":"梓官區"},{"zip":"827","name":"彌陀區"},{"zip":"828","name":"永安區"},{"zip":"829","name":"湖內區"},{"zip":"830","name":"鳳山區"},{"zip":"831","name":"大寮區"},{"zip":"832","name":"林園區"},{"zip":"833","name":"鳥松區"},{"zip":"840","name":"大樹區"},{"zip":"842","name":"旗山區"},{"zip":"843","name":"美濃區"},{"zip":"844","name":"六龜區"},{"zip":"845","name":"內門區"},{"zip":"846","name":"杉林區"},{"zip":"847","name":"甲仙區"},{"zip":"848","name":"桃源區"},{"zip":"849","name":"那瑪夏區"},{"zip":"851","name":"茂林區"},{"zip":"852","name":"茄萣區"}],"name":"高雄市"},{"districts":[{"zip":"817","name":"東沙群島"},{"zip":"819","name":"南沙群島"}],"name":"南海島"},{"districts":[{"zip":"880","name":"馬公市"},{"zip":"881","name":"西嶼鄉"},{"zip":"882","name":"望安鄉"},{"zip":"883","name":"七美鄉"},{"zip":"884","name":"白沙鄉"},{"zip":"885","name":"湖西鄉"}],"name":"澎湖縣"},{"districts":[{"zip":"890","name":"金沙鎮"},{"zip":"891","name":"金湖鎮"},{"zip":"892","name":"金寧鄉"},{"zip":"893","name":"金城鎮"},{"zip":"894","name":"烈嶼鄉"},{"zip":"896","name":"烏坵鄉"}],"name":"金門縣"},{"districts":[{"zip":"900","name":"屏東市"},{"zip":"901","name":"三地門鄉"},{"zip":"902","name":"霧台鄉"},{"zip":"903","name":"瑪家鄉"},{"zip":"904","name":"九如鄉"},{"zip":"905","name":"里港鄉"},{"zip":"906","name":"高樹鄉"},{"zip":"907","name":"鹽埔鄉"},{"zip":"908","name":"長治鄉"},{"zip":"909","name":"麟洛鄉"},{"zip":"911","name":"竹田鄉"},{"zip":"912","name":"內埔鄉"},{"zip":"913","name":"萬丹鄉"},{"zip":"920","name":"潮州鎮"},{"zip":"921","name":"泰武鄉"},{"zip":"922","name":"來義鄉"},{"zip":"923","name":"萬巒鄉"},{"zip":"924","name":"崁頂鄉"},{"zip":"925","name":"新埤鄉"},{"zip":"926","name":"南州鄉"},{"zip":"927","name":"林邊鄉"},{"zip":"928","name":"東港鎮"},{"zip":"929","name":"琉球鄉"},{"zip":"931","name":"佳冬鄉"},{"zip":"932","name":"新園鄉"},{"zip":"940","name":"枋寮鄉"},{"zip":"941","name":"枋山鄉"},{"zip":"942","name":"春日鄉"},{"zip":"943","name":"獅子鄉"},{"zip":"944","name":"車城鄉"},{"zip":"945","name":"牡丹鄉"},{"zip":"946","name":"恆春鎮"},{"zip":"947","name":"滿州鄉"}],"name":"屏東縣"},{"districts":[{"zip":"950","name":"台東市"},{"zip":"951","name":"綠島鄉"},{"zip":"952","name":"蘭嶼鄉"},{"zip":"953","name":"延平鄉"},{"zip":"954","name":"卑南鄉"},{"zip":"955","name":"鹿野鄉"},{"zip":"956","name":"關山鎮"},{"zip":"957","name":"海端鄉"},{"zip":"958","name":"池上鄉"},{"zip":"959","name":"東河鄉"},{"zip":"961","name":"成功鎮"},{"zip":"962","name":"長濱鄉"},{"zip":"963","name":"太麻里鄉"},{"zip":"964","name":"金峰鄉"},{"zip":"965","name":"大武鄉"},{"zip":"966","name":"達仁鄉"}],"name":"台東縣"},{"districts":[{"zip":"970","name":"花蓮市"},{"zip":"971","name":"新城鄉"},{"zip":"972","name":"秀林鄉"},{"zip":"973","name":"吉安鄉"},{"zip":"974","name":"壽豐鄉"},{"zip":"975","name":"鳳林鎮"},{"zip":"976","name":"光復鄉"},{"zip":"977","name":"豐濱鄉"},{"zip":"978","name":"瑞穗鄉"},{"zip":"979","name":"萬榮鄉"},{"zip":"981","name":"玉里鎮"},{"zip":"982","name":"卓溪鄉"},{"zip":"983","name":"富里鄉"}],"name":"花蓮縣"}]')



# 抓取時間
dt1 = datetime.utcnow().replace(tzinfo=timezone.utc)
now_time = dt1.astimezone(timezone(timedelta(hours=8))) # 轉換時區 -> 東八區
now_today = str(now_time.strftime("%Y-%m-%d"))
now_hour = str(now_time.strftime("%H:%M"))

window = tk.Tk()
window.title(VERSION + ' ' + str(now_today))
window.geometry('400x370')

hosts = [
 'ntp.ntu.edu.tw'
 ,'tock.stdtime.gov.tw'
 ,'watch.stdtime.gov.tw'
 ,'time.stdtime.gov.tw'
 ,'clock.stdtime.gov.tw'
 ,'tick.stdtime.gov.tw'
 ,'0.tw.pool.ntp.org'
 ,'1.tw.pool.ntp.org'
 ,'2.tw.pool.ntp.org'
 ,'3.tw.pool.ntp.org'
 ,'0.pool.ntp.org'
 ,'1.pool.ntp.org'
 ,'2.pool.ntp.org'
 ,'3.pool.ntp.org'
]

labVerify  = tk.Label(window, text = '請輸入驗證碼:', justify=tk.RIGHT, width=50)
labVerify.place(x=10, y=9, width=100, height=20)
varVerify = tk.StringVar()
varVerify.set('')

if os.path.isfile('config.txt'):
    f = open(r'config.txt')
    varVerify.set(f.readlines())
    f.close()

entVerify = tk.Entry(window, width = 120, textvariable = varVerify)
entVerify.place(x=110, y=9, width=230, height=20)

def ntp_client():
    global hosts

    # 打亂 hosts 順序以免回傳受到延遲

    random.shuffle(hosts)

    #建立例項，NTPClient()是一個類
    t = ntplib.NTPClient()
    for host in hosts:
        try:
            #ntp server可以填寫主機和域名，建議用域名
            #預設埠為ntp， 版本為2， 超時為5s
            #作用：查詢 NTP 伺服器，並返回物件
            r = t.request(host , port='ntp', version=4, timeout=3)
            if r:
                break
        except:
            pass

    try:
        t = r.tx_time
        #使用datetime模組,格式化：x年x月x日 時:分:秒.毫秒
        _date,_time = str(datetime.fromtimestamp(t))[:22].split(' ')
        # print("調整前時間是：", datetime.now())
        os.system('date {} && time {}'.format(_date, _time))
        # print("調整後時間是：", datetime.now())
    except:
        addInfo('無法取得伺服器時間，請重新啟動程式')
        return

    return _date,_time.split(':')[0] + ':' +_time.split(':')[1]

def VerifyCode():
    global VERIFYPASS,VERIFYCOUNT, comType, now_today, now_hour

    if len(entVerify.get()) != 128:
        if VERIFYCOUNT <=3:
            addInfo('輸入錯誤, 嘗試次數' + str(VERIFYCOUNT) + ' 次, 3次後關閉程式')
            VERIFYCOUNT = VERIFYCOUNT + 1
            return
        else:
            sys.exit()

    addInfo('檢查日期及時間中...')
    check_date = None
    check_time = None
    check_date, check_time = ntp_client()
    if check_date != now_today or check_time != now_hour:
        addInfo('讀取時間異常，3秒後關閉')
        time.sleep(1)
        addInfo('讀取時間異常，2秒後關閉')
        time.sleep(1)
        addInfo('讀取時間異常，1秒後關閉')
        time.sleep(1)
        sys.exit()

    # 初始化firestore ===========================================================
    pass_doc_id = []
    pass_list = []
    pass_name = []
    login_count = []
    scope_list = []
    news_list = []
    try:
        db = firestore.client()

        passcode_list_docs = db.collection(u'passcode_list').stream()
        for doc in passcode_list_docs:
            # print(f'{doc.id} => {doc.to_dict()}')
            pass_doc_id.append(doc.id)
            pass_list.append(doc.to_dict()['code'])
            pass_name.append(doc.to_dict()['name'])
            login_count.append(doc.to_dict()['login_count'])

        USERNAME = pass_name[pass_list.index(entVerify.get())]

        user_list_docs = db.collection(u'user_list').where(u'name', u'==', USERNAME).stream()
        for doc in user_list_docs:
            # 試用結束時間
            endTime = doc.to_dict()['end_date']
            # 開放權限
            scope_list = doc.to_dict()['rent_scope'] + doc.to_dict()['sale_scope']

        news_list_docs = db.collection(u'news').stream()
        for doc in news_list_docs:
            # 最新消息
            news_list = doc.to_dict()['content']
    except:
        addInfo('驗證失敗...，請提供相關畫面聯繫作者協助')
        return

    if len(pass_list) <= 0:
        addInfo('您的驗證碼尚未開通，請提供相關畫面請聯繫作者協助')
    if len(scope_list) <= 0:
        addInfo('您的選單功能尚未開通，請提供相關畫面請聯繫作者協助')

    # ==============================================================================

    comType = tt.Combobox(window, width=50, values=scope_list)
    comType.place(x=110, y=31, width=150, height=20)
    comType.bind("<<ComboboxSelected>>", callbackFunc3)

    # 建立 SHA1 物件
    s = hashlib.sha512()

    s.update(get_mac_address().encode("utf-8") + encode("utf-8"))
    h = s.hexdigest()

    if h == entVerify.get() and entVerify.get() in pass_list:
        if datetime.timestamp(endTime) - datetime.timestamp(now_time) > 0:
            addInfo('成功! 您已申請' + str(pass_name.count(USERNAME)) + '電腦, 請接著選擇類型縣市區域')
            VERIFYPASS = True
            f = open("config.txt","w+")
            f.write(entVerify.get())
            f.close()
            passcode_list_doc = db.collection(u'passcode_list').document(pass_doc_id[int(pass_list.index(entVerify.get()))])

            passcode_list_doc.update({
                u'last_login': now_time.strftime("%Y/%m/%d, %H:%M:%S"),
                u'login_count': int(login_count[int(pass_list.index(entVerify.get()))]) + 1
            })

            if (len(news_list) > 0):
                for news_data in news_list:
                    addInfo(news_data)

            btnVerify = tk.Button(window, text='OK', width=100, command=VerifyCode, state=tk.DISABLED)
            btnVerify.place(x=345, y=9, width=25, height=20)
        else:
            addInfo('超過試用期5秒後程式自動關閉')
            time.sleep(1)
            addInfo('超過試用期4秒後程式自動關閉')
            time.sleep(1)
            addInfo('超過試用期3秒後程式自動關閉')
            time.sleep(1)
            addInfo('超過試用期2秒後程式自動關閉')
            time.sleep(1)
            addInfo('超過試用期1秒後程式自動關閉')
            time.sleep(1)
            sys.exit()
    else:
        if VERIFYCOUNT <=3:
            # 用VPN可能導致失敗
            addInfo('輸入錯誤, 嘗試次數' + str(VERIFYCOUNT) + ' 次, 3次後關閉程式')
            VERIFYCOUNT = VERIFYCOUNT + 1
        else:
            sys.exit()
btnVerify = tk.Button(window, text='OK', width=100, command=VerifyCode)
btnVerify.place(x=345, y=9, width=25, height=20)

def callbackFunc3(event):
    global MAIN_URL
    if comType.get() == '租屋':
        MAIN_URL = "https://rent.591.com.tw/?kind=0&shType=host"
    if comType.get() == '店面出租':
        MAIN_URL = "https://business.591.com.tw/?type=1&kind=5"
    if comType.get() == '店面出售':
        MAIN_URL = "https://business.591.com.tw/?type=2&kind=5"
    if comType.get() == '辦公出租':
        MAIN_URL = "https://business.591.com.tw/?type=1&kind=6"
    if comType.get() == '辦公出售':
        MAIN_URL = "https://business.591.com.tw/?type=2&kind=6"
    if comType.get() == '住辦出租':
        MAIN_URL = "https://business.591.com.tw/?type=1&kind=12"
    if comType.get() == '住辦出售':
        MAIN_URL = "https://business.591.com.tw/?type=2&kind=12"
    if comType.get() == '廠房出租':
        MAIN_URL = "https://business.591.com.tw/?type=1&kind=7"
    if comType.get() == '廠房出售':
        MAIN_URL = "https://business.591.com.tw/?type=2&kind=7"
    if comType.get() == '土地出租':
        MAIN_URL = "https://business.591.com.tw/?type=1&kind=11"
    if comType.get() == '土地出售':
        MAIN_URL = "https://business.591.com.tw/?type=2&kind=11"
    if comType.get() == '中古屋':
        MAIN_URL = "https://sale.591.com.tw/"

    btnAdd = tk.Button(window, text='匯出檔案', width=40, state=tk.DISABLED)
    btnAdd.place(x=150, y=120, width=100, height=20)

labType = tk.Label(window, text = '類型:', justify=tk.RIGHT, width=50)
labType.place(x=10, y=31, width=100, height=20)



labCity = tk.Label(window, text = '縣市:', justify=tk.RIGHT, width=50)
labCity.place(x=10, y=52, width=100, height=20)

stdCity = (
 '台北市','新北市','桃園市','新竹市','新竹縣'
,'基隆市','宜蘭縣','台中市','彰化縣','苗栗縣'
,'雲林縣','高雄市','台南市','嘉義市','屏東縣'
,'嘉義縣','花蓮縣','台東縣','南投縣','澎湖縣'
,'金門縣','連江縣'
)

comCity = tt.Combobox(window, width=50, values=stdCity)
comCity.place(x=110, y=52, width=150, height=20)

def callbackFunc2(event):
    btnAdd = tk.Button(window, text='匯出檔案', width=40, state=tk.DISABLED)
    btnAdd.place(x=150, y=120, width=100, height=20)

def callbackFunc1(event):
    global stdArea, comArea
    stdArea = ('請選擇',)
    for item in Taiwan:
        if item['name'] == comCity.get():
            for AreaList in item['districts']:
                stdArea = stdArea + (AreaList['name'],)
                # print(stdArea)

    comArea = tt.Combobox(window, width=50, values=stdArea)
    comArea.place(x=110, y=74, width=150, height=20)
    comArea.bind("<<ComboboxSelected>>", callbackFunc2)

    btnAdd = tk.Button(window, text='匯出檔案', width=40, state=tk.DISABLED)
    btnAdd.place(x=150, y=120, width=100, height=20)

comCity.bind("<<ComboboxSelected>>", callbackFunc1)

stdArea = ('請選擇',)
labArea = tk.Label(window, text = '區域:', justify=tk.RIGHT, width=50)
labArea.place(x=10, y=74, width=100, height=20)

comArea = tt.Combobox(window, width=50, values=stdArea)
comArea.place(x=110, y=74, width=150, height=20)
comArea.bind("<<ComboboxSelected>>", callbackFunc2)

def getSearchURL():
    global sSection

    sUrlParaGet = ''
    if (comCity.get() == '台北市'):
        sUrlParaGet = '&region=1'
        if (comArea.get() == '中山區'): sSection = '1'
        if (comArea.get() == '大安區'): sSection = '5'
        if (comArea.get() == '信義區'): sSection = '7'
        if (comArea.get() == '內湖區'): sSection = '10'
        if (comArea.get() == '士林區'): sSection = '8'
        if (comArea.get() == '中正區'): sSection = '1'
        if (comArea.get() == '松山區'): sSection = '4'
        if (comArea.get() == '大同區'): sSection = '2'
        if (comArea.get() == '萬華區'): sSection = '6'
        if (comArea.get() == '北投區'): sSection = '9'
        if (comArea.get() == '文山區'): sSection = '12'
        if (comArea.get() == '南港區'): sSection = '11'

    if (comCity.get() == '新北市'):
        sUrlParaGet = '&region=3'
        if (comArea.get() == '板橋區'): sSection = '26'
        if (comArea.get() == '三重區'): sSection = '43'
        if (comArea.get() == '中和區'): sSection = '38'
        if (comArea.get() == '新莊區'): sSection = '44'
        if (comArea.get() == '淡水區'): sSection = '50'
        if (comArea.get() == '永和區'): sSection = '37'
        if (comArea.get() == '新店區'): sSection = '34'
        if (comArea.get() == '汐止區'): sSection = '27'
        if (comArea.get() == '蘆洲區'): sSection = '47'
        if (comArea.get() == '林口區'): sSection = '46'
        if (comArea.get() == '土城區'): sSection = '39'
        if (comArea.get() == '三峽區'): sSection = '40'
        if (comArea.get() == '樹林區'): sSection = '41'
        if (comArea.get() == '鶯歌區'): sSection = '42'
        if (comArea.get() == '泰山區'): sSection = '45'
        if (comArea.get() == '五股區'): sSection = '48'
        if (comArea.get() == '深坑區'): sSection = '28'
        if (comArea.get() == '八里區'): sSection = '49'
        if (comArea.get() == '三芝區'): sSection = '51'
        if (comArea.get() == '萬里區'): sSection = '20'
        if (comArea.get() == '瑞芳區'): sSection = '30'
        if (comArea.get() == '金山區'): sSection = '21'
        if (comArea.get() == '平溪區'): sSection = '31'
        if (comArea.get() == '貢寮區'): sSection = '33'
        if (comArea.get() == '石碇區'): sSection = '29'
        if (comArea.get() == '雙溪區'): sSection = '32'
        if (comArea.get() == '坪林區'): sSection = '35'
        if (comArea.get() == '烏來區'): sSection = '36'
        if (comArea.get() == '石門區'): sSection = '52'

    if (comCity.get() == '桃園市'):
        sUrlParaGet = '&region=6'
        if (comArea.get() == '桃園區'): sSection = '73'
        if (comArea.get() == '中壢區'): sSection = '67'
        if (comArea.get() == '蘆竹區'): sSection = '79'
        if (comArea.get() == '龜山區'): sSection = '74'
        if (comArea.get() == '八德區'): sSection = '75'
        if (comArea.get() == '大園區'): sSection = '78'
        if (comArea.get() == '平鎮區'): sSection = '68'
        if (comArea.get() == '楊梅區'): sSection = '70'
        if (comArea.get() == '觀音區'): sSection = '72'
        if (comArea.get() == '龍潭區'): sSection = '69'
        if (comArea.get() == '大溪區'): sSection = '76'
        if (comArea.get() == '新屋區'): sSection = '71'
        if (comArea.get() == '復興區'): sSection = '77'

    if (comCity.get() == '新竹市'):
        sUrlParaGet = '&region=4'
        if (comArea.get() == '東區'): sSection = '371'
        if (comArea.get() == '北區'): sSection = '372'
        if (comArea.get() == '香山區'): sSection = '370'

    if (comCity.get() == '新竹縣'):
        sUrlParaGet = '&region=5'
        if (comArea.get() == '竹北市'): sSection = '54'
        if (comArea.get() == '湖口鄉'): sSection = '55'
        if (comArea.get() == '新豐鄉'): sSection = '56'
        if (comArea.get() == '竹東鎮'): sSection = '61'
        if (comArea.get() == '寶山鄉'): sSection = '60'
        if (comArea.get() == '新埔鎮'): sSection = '57'
        if (comArea.get() == '芎林鄉'): sSection = '59'
        if (comArea.get() == '關西鎮'): sSection = '58'
        if (comArea.get() == '五峰鄉'): sSection = '62'
        if (comArea.get() == '橫山鄉'): sSection = '63'
        if (comArea.get() == '尖石鄉'): sSection = '64'
        if (comArea.get() == '北埔鄉'): sSection = '65'
        if (comArea.get() == '峨嵋鄉'): sSection = '66'

    if (comCity.get() == '基隆市'):
        sUrlParaGet = '&region=2'
        if (comArea.get() == '中正區'): sSection = '15'
        if (comArea.get() == '仁愛區'): sSection = '13'
        if (comArea.get() == '信義區'): sSection = '14'
        if (comArea.get() == '安樂區'): sSection = '17'
        if (comArea.get() == '中山區'): sSection = '16'
        if (comArea.get() == '七堵區'): sSection = '19'
        if (comArea.get() == '暖暖區'): sSection = '18'

    if (comCity.get() == '宜蘭縣'):
        sUrlParaGet = '&region=21'
        if (comArea.get() == '宜蘭市'): sSection = '328'
        if (comArea.get() == '羅東鎮'): sSection = '333'
        if (comArea.get() == '礁溪鄉'): sSection = '330'
        if (comArea.get() == '頭城鎮'): sSection = '329'
        if (comArea.get() == '蘇澳鎮'): sSection = '338'
        if (comArea.get() == '五結鄉'): sSection = '336'
        if (comArea.get() == '冬山鄉'): sSection = '337'
        if (comArea.get() == '三星鄉'): sSection = '334'
        if (comArea.get() == '員山鄉'): sSection = '332'
        if (comArea.get() == '壯圍鄉'): sSection = '331'
        if (comArea.get() == '大同鄉'): sSection = '335'
        if (comArea.get() == '南澳鄉'): sSection = '339'

    if (comCity.get() == '台中市'):
        sUrlParaGet = '&region=8'
        if (comArea.get() == '西屯區'): sSection = '104'
        if (comArea.get() == '北區'): sSection = '102'
        if (comArea.get() == '北屯區'): sSection = '103'
        if (comArea.get() == '西區'): sSection = '101'
        if (comArea.get() == '南區'): sSection = '100'
        if (comArea.get() == '南屯區'): sSection = '105'
        if (comArea.get() == '大里區'): sSection = '107'
        if (comArea.get() == '東區'): sSection = '99'
        if (comArea.get() == '龍井區'): sSection = '121'
        if (comArea.get() == '中區'): sSection = '98'
        if (comArea.get() == '太平區'): sSection = '106'
        if (comArea.get() == '沙鹿區'): sSection = '120'
        if (comArea.get() == '大雅區'): sSection = '117'
        if (comArea.get() == '潭子區'): sSection = '116'
        if (comArea.get() == '豐原區'): sSection = '110'
        if (comArea.get() == '后里區'): sSection = '111'
        if (comArea.get() == '烏日區'): sSection = '109'
        if (comArea.get() == '清水區'): sSection = '123'
        if (comArea.get() == '梧棲區'): sSection = '122'
        if (comArea.get() == '霧峰區'): sSection = '108'
        if (comArea.get() == '大甲區'): sSection = '124'
        if (comArea.get() == '神岡區'): sSection = '118'
        if (comArea.get() == '大肚區'): sSection = '119'
        if (comArea.get() == '外埔區'): sSection = '125'
        if (comArea.get() == '東勢區'): sSection = '113'
        if (comArea.get() == '和平區'): sSection = '114'
        if (comArea.get() == '大安區'): sSection = '126'
        if (comArea.get() == '石岡區'): sSection = '112'
        if (comArea.get() == '新社區'): sSection = '115'

    if (comCity.get() == '彰化縣'):
        sUrlParaGet = '&region=10'
        if (comArea.get() == '彰化市'): sSection = '127'
        if (comArea.get() == '員林市'): sSection = '136'
        if (comArea.get() == '鹿港鎮'): sSection = '131'
        if (comArea.get() == '大村鄉'): sSection = '141'
        if (comArea.get() == '和美鎮'): sSection = '134'
        if (comArea.get() == '溪湖鎮'): sSection = '140'
        if (comArea.get() == '北斗鎮'): sSection = '144'
        if (comArea.get() == '伸港鄉'): sSection = '135'
        if (comArea.get() == '秀水鄉'): sSection = '130'
        if (comArea.get() == '田中鎮'): sSection = '143'
        if (comArea.get() == '溪州鄉'): sSection = '147'
        if (comArea.get() == '二林鎮'): sSection = '149'
        if (comArea.get() == '福興鄉'): sSection = '132'
        if (comArea.get() == '花壇鄉'): sSection = '129'
        if (comArea.get() == '埤頭鄉'): sSection = '146'
        if (comArea.get() == '永靖鄉'): sSection = '138'
        if (comArea.get() == '社頭鄉'): sSection = '137'
        if (comArea.get() == '線西鄉'): sSection = '133'
        if (comArea.get() == '埔鹽鄉'): sSection = '142'
        if (comArea.get() == '埔心鄉'): sSection = '139'
        if (comArea.get() == '芳苑鄉'): sSection = '151'
        if (comArea.get() == '二水鄉'): sSection = '152'
        if (comArea.get() == '芬園鄉'): sSection = '128'
        if (comArea.get() == '田尾鄉'): sSection = '145'
        if (comArea.get() == '竹塘鄉'): sSection = '148'
        if (comArea.get() == '大城鄉'): sSection = '150'

    if (comCity.get() == '苗栗縣'):
        sUrlParaGet = '&region=7'
        if (comArea.get() == '竹南鎮'): sSection = '80'
        if (comArea.get() == '頭份市'): sSection = '81'
        if (comArea.get() == '苗栗市'): sSection = '88'
        if (comArea.get() == '後龍鎮'): sSection = '85'
        if (comArea.get() == '苑裡鎮'): sSection = '87'
        if (comArea.get() == '銅鑼鄉'): sSection = '94'
        if (comArea.get() == '公館鄉'): sSection = '91'
        if (comArea.get() == '造橋鄉'): sSection = '89'
        if (comArea.get() == '通霄鎮'): sSection = '86'
        if (comArea.get() == '三義鄉'): sSection = '95'
        if (comArea.get() == '大湖鄉'): sSection = '92'
        if (comArea.get() == '卓蘭鎮'): sSection = '97'
        if (comArea.get() == '三灣鄉'): sSection = '82'
        if (comArea.get() == '南庄鄉'): sSection = '83'
        if (comArea.get() == '獅潭鄉'): sSection = '84'
        if (comArea.get() == '頭屋鄉'): sSection = '90'
        if (comArea.get() == '泰安鄉'): sSection = '93'
        if (comArea.get() == '西湖鄉'): sSection = '96'


    if (comCity.get() == '雲林縣'):
        sUrlParaGet = '&region=14'
        if (comArea.get() == '斗六市'): sSection = '194'
        if (comArea.get() == '虎尾鎮'): sSection = '187'
        if (comArea.get() == '斗南鎮'): sSection = '185'
        if (comArea.get() == '北港鎮'): sSection = '200'
        if (comArea.get() == '西螺鎮'): sSection = '198'
        if (comArea.get() == '麥寮鄉'): sSection = '193'
        if (comArea.get() == '臺西鄉'): sSection = '191'
        if (comArea.get() == '土庫鎮'): sSection = '188'
        if (comArea.get() == '東勢鄉'): sSection = '190'
        if (comArea.get() == '崙背鄉'): sSection = '192'
        if (comArea.get() == '口湖鄉'): sSection = '202'
        if (comArea.get() == '古坑鄉'): sSection = '196'
        if (comArea.get() == '元長鄉'): sSection = '204'
        if (comArea.get() == '褒忠鄉'): sSection = '189'
        if (comArea.get() == '二崙鄉'): sSection = '199'
        if (comArea.get() == '大埤鄉'): sSection = '186'
        if (comArea.get() == '莿桐鄉'): sSection = '197'
        if (comArea.get() == '水林鄉'): sSection = '201'
        if (comArea.get() == '林內鄉'): sSection = '195'
        if (comArea.get() == '四湖鄉'): sSection = '203'

    if (comCity.get() == '南投縣'):
        sUrlParaGet = '&region=11'
        if (comArea.get() == '南投市'): sSection = '153'
        if (comArea.get() == '草屯鎮'): sSection = '155'
        if (comArea.get() == '埔里鎮'): sSection = '157'
        if (comArea.get() == '竹山鎮'): sSection = '164'
        if (comArea.get() == '鹿谷鄉'): sSection = '165'
        if (comArea.get() == '魚池鄉'): sSection = '162'
        if (comArea.get() == '集集鎮'): sSection = '160'
        if (comArea.get() == '水里鄉'): sSection = '161'
        if (comArea.get() == '名間鄉'): sSection = '159'
        if (comArea.get() == '國姓鄉'): sSection = '156'
        if (comArea.get() == '中寮鄉'): sSection = '154'
        if (comArea.get() == '仁愛鄉'): sSection = '158'
        if (comArea.get() == '信義鄉'): sSection = '163'

    if (comCity.get() == '高雄市'):
        sUrlParaGet = '&region=17'
        if (comArea.get() == '三民區'): sSection = '250'
        if (comArea.get() == '苓雅區'): sSection = '245'
        if (comArea.get() == '左營區'): sSection = '253'
        if (comArea.get() == '前鎮區'): sSection = '249'
        if (comArea.get() == '楠梓區'): sSection = '251'
        if (comArea.get() == '鼓山區'): sSection = '247'
        if (comArea.get() == '鳳山區'): sSection = '268'
        if (comArea.get() == '新興區'): sSection = '243'
        if (comArea.get() == '前金區'): sSection = '244'
        if (comArea.get() == '小港區'): sSection = '252'
        if (comArea.get() == '鹽埕區'): sSection = '246'
        if (comArea.get() == '大社區'): sSection = '255'
        if (comArea.get() == '岡山區'): sSection = '258'
        if (comArea.get() == '鳥松區'): sSection = '271'
        if (comArea.get() == '仁武區'): sSection = '254'
        if (comArea.get() == '大寮區'): sSection = '269'
        if (comArea.get() == '橋頭區'): sSection = '263'
        if (comArea.get() == '路竹區'): sSection = '259'
        if (comArea.get() == '燕巢區'): sSection = '262'
        if (comArea.get() == '林園區'): sSection = '270'
        if (comArea.get() == '湖內區'): sSection = '267'
        if (comArea.get() == '大樹區'): sSection = '272'
        if (comArea.get() == '旗津區'): sSection = '248'
        if (comArea.get() == '梓官區'): sSection = '264'
        if (comArea.get() == '旗山區'): sSection = '273'
        if (comArea.get() == '阿蓮區'): sSection = '260'
        if (comArea.get() == '永安區'): sSection = '266'
        if (comArea.get() == '茄萣區'): sSection = '282'
        if (comArea.get() == '甲仙區'): sSection = '278'
        if (comArea.get() == '美濃區'): sSection = '274'
        if (comArea.get() == '田寮區'): sSection = '261'
        if (comArea.get() == '彌陀區'): sSection = '265'
        if (comArea.get() == '六龜區'): sSection = '275'
        if (comArea.get() == '內門區'): sSection = '276'
        if (comArea.get() == '杉林區'): sSection = '277'
        if (comArea.get() == '桃源區'): sSection = '279'
        if (comArea.get() == '那瑪夏區'): sSection = '280'
        if (comArea.get() == '茂林區'): sSection = '281'

    if (comCity.get() == '台南市'):
        sUrlParaGet = '&region=15'
        if (comArea.get() == '永康區'): sSection = '312'
        if (comArea.get() == '東區'): sSection = '206'
        if (comArea.get() == '北區'): sSection = '209'
        if (comArea.get() == '中西區'): sSection = '208'
        if (comArea.get() == '安平區'): sSection = '210'
        if (comArea.get() == '南區'): sSection = '207'
        if (comArea.get() == '仁德區'): sSection = '219'
        if (comArea.get() == '新市區'): sSection = '241'
        if (comArea.get() == '善化區'): sSection = '238'
        if (comArea.get() == '安南區'): sSection = '211'
        if (comArea.get() == '新營區'): sSection = '230'
        if (comArea.get() == '歸仁區'): sSection = '213'
        if (comArea.get() == '佳里區'): sSection = '224'
        if (comArea.get() == '麻豆區'): sSection = '223'
        if (comArea.get() == '柳營區'): sSection = '236'
        if (comArea.get() == '安定區'): sSection = '242'
        if (comArea.get() == '官田區'): sSection = '222'
        if (comArea.get() == '新化區'): sSection = '214'
        if (comArea.get() == '玉井區'): sSection = '216'
        if (comArea.get() == '鹽水區'): sSection = '237'
        if (comArea.get() == '白河區'): sSection = '232'
        if (comArea.get() == '六甲區'): sSection = '234'
        if (comArea.get() == '後壁區'): sSection = '231'
        if (comArea.get() == '西港區'): sSection = '225'
        if (comArea.get() == '關廟區'): sSection = '220'
        if (comArea.get() == '山上區'): sSection = '240'
        if (comArea.get() == '楠西區'): sSection = '217'
        if (comArea.get() == '左鎮區'): sSection = '215'
        if (comArea.get() == '南化區'): sSection = '218'
        if (comArea.get() == '龍崎區'): sSection = '221'
        if (comArea.get() == '七股區'): sSection = '226'
        if (comArea.get() == '將軍區'): sSection = '227'
        if (comArea.get() == '學甲區'): sSection = '228'
        if (comArea.get() == '北門區'): sSection = '229'
        if (comArea.get() == '東山區'): sSection = '233'
        if (comArea.get() == '下營區'): sSection = '235'
        if (comArea.get() == '大內區'): sSection = '239'

    if (comCity.get() == '嘉義市'):
        sUrlParaGet = '&region=12'
        if (comArea.get() == '西區'): sSection = '373'
        if (comArea.get() == '東區'): sSection = '374'

    if (comCity.get() == '屏東縣'):
        sUrlParaGet = '&region=19'
        if (comArea.get() == '屏東市'): sSection = '295'
        if (comArea.get() == '東港鎮'): sSection = '316'
        if (comArea.get() == '內埔鄉'): sSection = '306'
        if (comArea.get() == '潮州鎮'): sSection = '308'
        if (comArea.get() == '恆春鎮'): sSection = '326'
        if (comArea.get() == '長治鄉'): sSection = '303'
        if (comArea.get() == '鹽埔鄉'): sSection = '302'
        if (comArea.get() == '萬丹鄉'): sSection = '307'
        if (comArea.get() == '枋寮鄉'): sSection = '320'
        if (comArea.get() == '南州鄉'): sSection = '314'
        if (comArea.get() == '車城鄉'): sSection = '324'
        if (comArea.get() == '萬巒鄉'): sSection = '311'
        if (comArea.get() == '里港鄉'): sSection = '300'
        if (comArea.get() == '新園鄉'): sSection = '319'
        if (comArea.get() == '滿州鄉'): sSection = '327'
        if (comArea.get() == '高樹鄉'): sSection = '301'
        if (comArea.get() == '崁頂鄉'): sSection = '312'
        if (comArea.get() == '琉球鄉'): sSection = '317'
        if (comArea.get() == '三地門鄉'): sSection = '296'
        if (comArea.get() == '麟洛鄉'): sSection = '304'
        if (comArea.get() == '九如鄉'): sSection = '299'
        if (comArea.get() == '林邊鄉'): sSection = '315'
        if (comArea.get() == '霧臺鄉'): sSection = '297'
        if (comArea.get() == '瑪家鄉'): sSection = '298'
        if (comArea.get() == '竹田鄉'): sSection = '305'
        if (comArea.get() == '泰武鄉'): sSection = '309'
        if (comArea.get() == '來義鄉'): sSection = '310'
        if (comArea.get() == '新埤鄉'): sSection = '313'
        if (comArea.get() == '佳冬鄉'): sSection = '318'
        if (comArea.get() == '枋山鄉'): sSection = '321'
        if (comArea.get() == '春日鄉'): sSection = '322'
        if (comArea.get() == '獅子鄉'): sSection = '323'
        if (comArea.get() == '牡丹鄉'): sSection = '325'

    if (comCity.get() == '嘉義縣'):
        sUrlParaGet = '&region=13'
        if (comArea.get() == '民雄鄉'): sSection = '180'
        if (comArea.get() == '太保市'): sSection = '175'
        if (comArea.get() == '朴子市'): sSection = '176'
        if (comArea.get() == '大林鎮'): sSection = '181'
        if (comArea.get() == '梅山鄉'): sSection = '168'
        if (comArea.get() == '水上鄉'): sSection = '173'
        if (comArea.get() == '中埔鄉'): sSection = '171'
        if (comArea.get() == '新港鄉'): sSection = '179'
        if (comArea.get() == '義竹鄉'): sSection = '183'
        if (comArea.get() == '竹崎鄉'): sSection = '169'
        if (comArea.get() == '番路鄉'): sSection = '167'
        if (comArea.get() == '阿里山鄉'): sSection = '170'
        if (comArea.get() == '大埔鄉'): sSection = '172'
        if (comArea.get() == '鹿草鄉'): sSection = '174'
        if (comArea.get() == '東石鄉'): sSection = '177'
        if (comArea.get() == '六腳鄉'): sSection = '178'
        if (comArea.get() == '溪口鄉'): sSection = '182'
        if (comArea.get() == '布袋鎮'): sSection = '184'


    if (comCity.get() == '花蓮縣'):
        sUrlParaGet = '&region=23'
        if (comArea.get() == '花蓮市'): sSection = '357'
        if (comArea.get() == '吉安鄉'): sSection = '360'
        if (comArea.get() == '壽豐鄉'): sSection = '361'
        if (comArea.get() == '新城鄉'): sSection = '358'
        if (comArea.get() == '玉里鎮'): sSection = '367'
        if (comArea.get() == '瑞穗鄉'): sSection = '365'
        if (comArea.get() == '鳳林鎮'): sSection = '362'
        if (comArea.get() == '秀林鄉'): sSection = '359'
        if (comArea.get() == '光復鄉'): sSection = '363'
        if (comArea.get() == '豐濱鄉'): sSection = '364'
        if (comArea.get() == '萬榮鄉'): sSection = '366'
        if (comArea.get() == '卓溪鄉'): sSection = '368'
        if (comArea.get() == '富里鄉'): sSection = '369'

    if (comCity.get() == '台東縣'):
        sUrlParaGet = '&region=22'
        if (comArea.get() == '台東市'): sSection = '341'
        if (comArea.get() == '成功鎮'): sSection = '351'
        if (comArea.get() == '卑南鄉'): sSection = '345'
        if (comArea.get() == '池上鄉'): sSection = '349'
        if (comArea.get() == '東河鄉'): sSection = '350'
        if (comArea.get() == '太麻里鄉'): sSection = '353'
        if (comArea.get() == '綠島鄉'): sSection = '342'
        if (comArea.get() == '蘭嶼鄉'): sSection = '343'
        if (comArea.get() == '延平鄉'): sSection = '344'
        if (comArea.get() == '鹿野鄉'): sSection = '346'
        if (comArea.get() == '關山鎮'): sSection = '347'
        if (comArea.get() == '海端鄉'): sSection = '348'
        if (comArea.get() == '長濱鄉'): sSection = '352'
        if (comArea.get() == '金峰鄉'): sSection = '354'
        if (comArea.get() == '大武鄉'): sSection = '355'
        if (comArea.get() == '達仁鄉'): sSection = '356'

    if (comCity.get() == '金門縣'):
        sUrlParaGet = '&region=25'
        if (comArea.get() == '金寧鄉'): sSection = '291'
        if (comArea.get() == '金湖鎮'): sSection = '290'
        if (comArea.get() == '金城鎮'): sSection = '292'
        if (comArea.get() == '金沙鎮'): sSection = '289'
        if (comArea.get() == '烈嶼鄉'): sSection = '293'
        if (comArea.get() == '烏坵鄉'): sSection = '294'

    if (comCity.get() == '澎湖縣'):
        sUrlParaGet = '&region=24'
        if (comArea.get() == '馬公市'): sSection = '283'
        if (comArea.get() == '西嶼鄉'): sSection = '284'
        if (comArea.get() == '望安鄉'): sSection = '285'
        if (comArea.get() == '七美鄉'): sSection = '286'
        if (comArea.get() == '白沙鄉'): sSection = '287'
        if (comArea.get() == '湖西鄉'): sSection = '288'

    if (comCity.get() == '連江縣'):
        sUrlParaGet = '&region=26'
        if (comArea.get() == '東引鄉'): sSection = '25'
        if (comArea.get() == '南竿鄉'): sSection = '22'
        if (comArea.get() == '北竿鄉'): sSection = '23'
        if (comArea.get() == '莒光鄉'): sSection = '24'
        if (comArea.get() == '東沙'): sSection = '256'
        if (comArea.get() == '南沙'): sSection = '257'

    sUrlParaGet += '&section=' + sSection

    return sUrlParaGet

def setChrome():
    global totalpages, MAIN_URL, browser, sSection, options, varDownloadPage, SHOWCHROME

    try:
        if comType.get() in SEARCH_A:
            browser = webdriver.Chrome(executable_path=ChromeDriverManager().install(),chrome_options=options)
            if SHOWCHROME != True:
                browser.set_window_position(-10000,0)
            browser.get(MAIN_URL + getSearchURL())

            #關閉選取地區pop-up 否則無法點選下一頁
            browser.find_element_by_css_selector('dd[google-data-stat="首頁_縣市選擇_' + comCity.get() + '"]').click()
            time.sleep(3)
            #輸入 ESC 關閉google 提示，否則無法點選
            try:
                browser.find_element_by_class_name('pageNext').send_keys(Keys.ESCAPE) #ECS鍵
                bs = BeautifulSoup(browser.page_source, 'html.parser')
                totalpages = int(int(bs.find('span', {'class':'TotalRecord'}).text.split(' ')[-2])/30) + 1
            except:
                totalpages = 1
                addInfo('running code:' + printLineFileFunc())

        if comType.get() in SEARCH_B:
            browser = webdriver.Chrome(executable_path=ChromeDriverManager().install(),chrome_options=options)
            if SHOWCHROME != True:
                browser.set_window_position(-10000,0)
            browser.get(MAIN_URL + getSearchURL())
            time.sleep(3)
            # 選擇縣市 ==============================================================
            try:
                browser.find_element_by_css_selector('span[google-data-stat="商用出租_列表_按縣市選擇"').click()
                time.sleep(3)
            except:
                print()

            try:
                browser.find_element_by_css_selector('span[google-data-stat="商用出售_列表_按縣市選擇"').click()
                time.sleep(3)
            except:
                print()

            # 選擇區域 ===============================================================
            try:
                browser.find_element_by_css_selector('a[google-data-stat="縣市選擇_大區域_' + comCity.get() + '"]').click()
                time.sleep(3)
            except:
                print()

            try:
                browser.find_element_by_css_selector('div[google-data-stat="新版出租_列表_鄉鎮更多"').click()
            except:
                print()

            try:
                browser.find_element_by_css_selector('div[google-data-stat="商用出租_列表_按鄉鎮選擇"').click()
            except:
                print()

            time.sleep(5)
            browser.find_element_by_css_selector('label[for="checktips' + sSection + '"').click()
            time.sleep(3)
            browser.find_element_by_css_selector('li[data-text="host"]').click()
            time.sleep(3)
            try:
                browser.find_element_by_class_name('pageNext').send_keys(Keys.ESCAPE) #ECS鍵
                bs = BeautifulSoup(browser.page_source, 'html.parser')
                # totalpages = int(int(bs.find('span', {'class':'TotalRecord'}).text.split(' ')[-2])/30) + 1
                totalpages = int(int(bs.find('span', {'class':'R'}).text)/30) + 1
            except:
                totalpages = 1
                addInfo('running code:' + printLineFileFunc())

        if comType.get() in SEARCH_C:
            browser = webdriver.Chrome(executable_path=ChromeDriverManager().install(),chrome_options=options)
            if SHOWCHROME != True:
                browser.set_window_position(-10000,0)
            browser.get(MAIN_URL)

            if browser.find_element_by_css_selector('div[class*="tips-popbox-img"]').is_displayed():
                time.sleep(3)
                browser.find_element_by_css_selector('div[class*="tips-popbox-img"]').click()

            if browser.find_element_by_css_selector('div[class*="accreditPop"]').is_displayed():
                time.sleep(3)
                browser.find_element_by_css_selector('div[class*="accreditPop"]').click()


            #選擇縣市
            browser.find_element_by_css_selector('div[class*="filter-region"]').click()
            time.sleep(3)
            listCity = browser.find_elements_by_xpath("//div[contains(@class,'region-list-item')]")
            for oCity in listCity:
                if (oCity.text == comCity.get()):
                    oCity.click()
            time.sleep(3)

            #選擇區域
            listArea = browser.find_elements_by_xpath("//a[contains(@class,'section-list-item-link')]")
            # listArea = browser.find_element_by_css_selector('a[class*="section-list-item-link"]')
            for oArea in listArea:
                if (oArea.text == comArea.get()):
                    oArea.click()

            time.sleep(3)
            browser.find_element_by_css_selector('div[data-id="host"]').click()
            time.sleep(3)
            try:
                bs = BeautifulSoup(browser.page_source, 'html.parser')
                listPage = bs.find('a', {'class':'pageNum-form'})
                totalpages = int(int(listPage.attrs['data-total'])/30) + 1
            except:
                addInfo('running code:' + printLineFileFunc())
                totalpages = 1
    except:
        addInfo('running code:' + printLineFileFunc())


    try:
        if totalpages == 1 and 'display: block;' in browser.find_element_by_css_selector('div[class*="noInfo clearfix"]').get_attribute("style").strip():
            addInfo('您搜尋的範圍查無資料，請重新查詢')
            return
    except:
        addInfo('running code:' + printLineFileFunc())


    if (totalpages > 50):
        addInfo('頁數讀取異常，請重新搜尋')
        browser.quit()
    else:
        addInfo('共' + str(totalpages) + '頁, 請按[匯出檔案]按鈕, 勿關閉Chrome')

    labGetPage = tk.Label(window, text = '輸入下載頁數:', justify=tk.RIGHT, width=50)
    labGetPage.place(x=10, y=100, width=100, height=20)

    varDownloadPage = tk.StringVar()
    # varDownloadPage.set(str(1))
    entDownloadPage = tk.Entry(window, width = 120, textvariable = varDownloadPage)
    entDownloadPage.place(x=110, y=100, width=150, height=20)

    labTotalPage = tk.Label(window, text = '共' + str(totalpages) +'頁', justify=tk.RIGHT, width=50)
    labTotalPage.place(x=270, y=100, width=50, height=20)

    btnAdd = tk.Button(window, text='匯出檔案', width=40, command=getEachPage)
    btnAdd.place(x=150, y=120, width=100, height=20)

# input 物件網址 撈取網頁資料 return 欄位
def getHouseData_A(url):
    global now_today

    request_url='https:'+str(url).strip()
    res=requests.get(request_url)

    if res.status_code == 200:
        bs=BeautifulSoup(res.text,'html.parser')
        #先宣告變數為NULL 若無撈到資料則寫入NULL
        addr=''
        price=''
        size=''
        floor=''
        room_type=''
        now_environment= ''
        form=''
        car=''

        # 利用 beautfiulsoup 的 find function 利用 css selector 定位 並撈出指定資料
        addr=bs.find('span',{'class':'addr'}).text
        price=bs.find('div',{'class':'price'}).text.strip()
        room_attrs=bs.find('ul',{'class':'attr'}).findAll('li')
        for attr in room_attrs:
            if attr.text.split('\xa0:\xa0\xa0')[0]=='坪數':
                size=attr.text.split('\xa0:\xa0\xa0')[1]
            if attr.text.split('\xa0:\xa0\xa0')[0]=='面積':
                size=attr.text.split('\xa0:\xa0\xa0')[1]
            elif attr.text.split('\xa0:\xa0\xa0')[0]=='樓層':
                floor=attr.text.split('\xa0:\xa0\xa0')[1]
            elif attr.text.split('\xa0:\xa0\xa0')[0]=='型態':
                room_type=attr.text.split('\xa0:\xa0\xa0')[1]
            elif attr.text.split('\xa0:\xa0\xa0')[0]=='類別':
                room_type=attr.text.split('\xa0:\xa0\xa0')[1]
            elif attr.text.split('\xa0:\xa0\xa0')[0]=='現況':
                # print(attr.text.split('\xa0:\xa0\xa0'))
                now_environment = attr.text.split('\xa0:\xa0\xa0')[1]

        owner=bs.find('div',{'class':'userInfo'}).find('i').text

        room_descriptions=bs.find('ul',{'class':'labelList-1'}).findAll('li')
        for description in room_descriptions:
            if description.text.split('：')[0]=='格局':
                form=description.text.split('：')[1].replace('有陽台非於政府免付費公開資料可查詢法定用途', '')
                form=form.replace('法定用途', '')
            if re.sub(r"\s+", "", description.text.split('：')[0])=='車位':
                car= car + ' ' + description.text.split('：')[1]
                car= car.replace('管理費', '')
                car= car.replace('最短租期', '')
                car= car.replace('性別', '')
                car= car.replace('要求', '')

        person_name=bs.find('div',{'class':'avatarRight'}).findAll('i')[0].text
        phone=str(bs.find('span',{'class':'num'}).text)
        phone= re.sub(r"\s+", "", phone)

        if len(phone) == 0 or len(phone) == 2:
            phone=str(bs.find('span',{'class':'dialPhoneNum'}).text)


            if len(phone) == 0 or len(phone) == 2:
                phone=bs.find('span',{'class':'num'}).find('img')

                # 圖片存放路徑
                path = './phone_img_' + now_today + '/'
                now_img_name = path + str(time.time()) + '.png'

                os.makedirs(path ,exist_ok=True)
                r=requests.get('http:' + str(phone["src"]), headers={'User-Agent': UserAgent().chrome})
                with open(now_img_name,'wb') as f:
                    # 將圖片下載下來
                    f.write(r.content)
                phone=''

        return addr,price,size,floor,room_type,now_environment,car,owner,phone
    else:
        print('link expired:', url, res.status_code)
        return 404, 404, 404, 404, 404, 404, 404

def getHouseData_B(url):

    global now_today
    request_url=str(url).strip()
    res=requests.get(request_url, headers={'User-Agent': UserAgent().chrome})
    if res.status_code == 200:

        bs=BeautifulSoup(res.text,'html.parser')
        #先宣告變數為NULL 若無撈到資料則寫入NULL
        addr=''
        price=''
        size=''
        floor=''
        room_type=''
        now_environment=''
        form=''
        car=''
        age=''

        # 地址物件
        listAddr = bs.findAll('span',{'class':'info-addr-value'})
        addr = listAddr[len(listAddr)-1].text
        price = bs.find('span',{'class':'info-price-num'}).text + '萬 (' + bs.find('div',{'class':'info-price-per'}).text + ')'


        infoFloor = bs.findAll('div',{'class':'info-floor-left'})
        for info in infoFloor:
            if '樓層' in info.text:
                floor = info.find('div',{'class':'info-floor-key'}).text
            if '屋齡' in info.text:
                age = '(' + info.find('div',{'class':'info-floor-key'}).text + ')'
            if '坪數' in info.text:
                size = info.find('div',{'class':'info-floor-key'}).text
            if '車位' in info.text:
                car = info.find('div',{'class':'info-floor-key'}).text
        # size = '('+ listFloor[1].text + ') ' + listFloor[2].text
        # car = listFloor[2].text
        size = age + size

        listDetail = bs.findAll('div',{'class':'info-addr-content'})
        for detail in listDetail:
            try:
                if '現況' in detail.text:
                    now_environment = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
                if '型態' in detail.text:
                    room_type = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
                if '樓層' in detail.text:
                    floor = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
            except:
                addInfo('running code:' + printLineFileFunc())

        sDetailHouse = bs.findAll('div',{'class':'detail-house-content'})
        for detailhouse in sDetailHouse:
            try:
                if '現況' in detailhouse.find('div',{'class':'detail-house-key'}).text:
                    now_environment = str(detailhouse.find('div',{'class':'detail-house-value'}).text).replace('None', '')
                if '型態' in detailhouse.find('div',{'class':'detail-house-key'}).text:
                    room_type = str(detailhouse.find('div',{'class':'detail-house-value'}).text).replace('None', '')
            except:
                addInfo('running code:' + printLineFileFunc())


        owner = bs.find('span',{'class':'info-span-name'}).text.replace('（', '')
        phone = str(bs.find('span',{'class':'info-host-word'}).text).strip()


        return addr,price,size,floor,room_type,now_environment,car,owner,phone
    else:
        print('link expired:', url, res.status_code)
        return 404, 404, 404, 404, 404, 404, 404

def getHouseData_C(url):
    global now_today

    request_url=str(url).strip()
    res=requests.get(request_url, headers={'User-Agent': UserAgent().chrome})
    if res.status_code == 200:

        bs=BeautifulSoup(res.text,'html.parser')
        #先宣告變數為NULL 若無撈到資料則寫入NULL
        addr=''
        price=''
        size=''
        floor=''
        room_type=''
        now_environment=''
        form=''
        car=''

        # 地址物件
        listAddr = bs.findAll('span',{'class':'info-addr-value'})
        floor = listAddr[0].text
        addr = listAddr[len(listAddr)-1].text

        price = bs.find('span',{'class':'info-price-num'}).text + '萬 (' + bs.find('div',{'class':'info-price-per'}).text + ')'


        listFloor = bs.findAll('div',{'class':'info-floor-key'})
        size = '('+ listFloor[1].text + ') ' + listFloor[2].text
        car = listFloor[2].text

        listDetail = bs.findAll('div',{'class':'detail-house-value'})
        room_type = listDetail[1].text
        now_environment = listDetail[0].text
        # print(44,room_type, now_environment)

        owner = bs.find('span',{'class':'info-span-name'}).text.replace('（', '')
        phone = str(bs.find('span',{'class':'info-host-word'}).text).strip()
        # print(55,owner, phone)

        return addr,price,size,floor,room_type,now_environment,car,owner,phone
    else:
        print('link expired:', url, res.status_code)
        return 404, 404, 404, 404, 404, 404, 404

def getHouseData_D(url):
    global now_today
    request_url=str(url).strip()
    res=requests.get(request_url, headers={'User-Agent': UserAgent().chrome})

    if res.status_code == 200:
        bs=BeautifulSoup(res.text,'html.parser')
        #先宣告變數為NULL 若無撈到資料則寫入NULL
        addr=''
        price=''
        size=''
        floor=''
        room_type=''
        now_environment=''
        form=''
        car=''
        age=''

        # 地址物件
        listAddr = bs.findAll('span',{'class':'info-addr-value'})
        price = bs.find('span',{'class':'info-price-num'}).text + '萬 (' + bs.find('div',{'class':'info-price-per'}).text + ')'
        infoFloor = bs.findAll('div',{'class':'info-floor-left'})
        for info in infoFloor:
            if (info.text is not None):
                if '樓層' in info.text:
                    floor = str(info.find('div',{'class':'info-floor-key'}).text).replace('None', '')
                if '屋齡' in info.text:
                    age = '(' + str(info.find('div',{'class':'info-floor-key'}).text).replace('None', '') + ')'
                if '坪數' in info.text:
                    size = str(info.find('div',{'class':'info-floor-key'}).text).replace('None', '')
                if '面積' in info.text:
                    size = str(info.find('div',{'class':'info-floor-key'}).text).replace('None', '')
                if '車位' in info.text:
                    car = str(info.find('div',{'class':'info-floor-key'}).text).replace('None', '')

        # size = '('+ listFloor[1].text + ') ' + listFloor[2].text
        # car = listFloor[2].text
        size = age + size
        listDetail = bs.findAll('div',{'class':'info-addr-content'})

        for detail in listDetail:
            try:
                if '現況' in detail.text:
                    now_environment = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
                if '型態' in detail.text:
                    room_type = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
                if '地址' in detail.text:
                    addr = str(detail.find('span',{'class':'info-addr-value'}).text).replace('None', '')
            except:
                addInfo('running code:' + printLineFileFunc())

        owner = bs.find('span',{'class':'info-span-name'}).text.replace('（', '')
        phone = str(bs.find('span',{'class':'info-host-word'}).text).strip()
        print(addr,price,size,floor,room_type,now_environment,car,owner,phone)

        return addr,price,size,floor,room_type,now_environment,car,owner,phone
    else:
        print('link expired:', url, res.status_code)
        return 404, 404, 404, 404, 404, 404, 404

def getEachPage():
    global totalpages, browser, df, varDownloadPage, now_today

    try:
        checkDown = int(varDownloadPage.get())
    except:
        addInfo('請輸入下載頁數')
        addInfo('running code:' + printLineFileFunc())
        return

    if int(varDownloadPage.get()) > totalpages:
        addInfo('您輸入[' + str(varDownloadPage.get()) +'] 大於最大頁數，取最大頁數抓取資料')
        checkDown = totalpages

    if int(varDownloadPage.get()) <= 0:
        addInfo('請輸入頁數')
        return

    window.after(1000, parserWeb(checkDown,df, now_today))

def parserWeb(checkDown,df, now_today):
    count_rows = 0
    for i in range(int(checkDown)):
        # addInfo('目前進度.... ' + str(i+1) + '/' + str(checkDown) + ' 頁')
        room_url_list=[] #存放網址list
        # print('browser.page_source', browser.page_source)
        bs = BeautifulSoup(browser.page_source, 'html.parser')
        if comType.get() in WEBTYPE_A:
            titles=bs.findAll('h3') # h3 放置物件的區塊
            for title in titles:
                room_url=title.find('a').get('href') # 每個物件的 url
                room_url_list.append(room_url)
        if comType.get() in WEBTYPE_B or comType.get() in WEBTYPE_D:
            titles1= bs.findAll('div', {'class':'j-house houseList-item clearfix z-hastag'})
            if (len(titles1) > 0):
                for title in titles1:
                    room_url_list.append('https://sale.591.com.tw/home/house/detail/2/' + title.attrs['data-bind'] + '.html')
            titles2=bs.findAll('ul', {'class':'listInfo clearfix j-house'})
            if (len(titles2) > 0):
                for title in titles2:
                    room_url_list.append('https://sale.591.com.tw/home/house/detail/2/' + title.attrs['data-bind'] + '.html')
        # ------------- GET data ------------- #
        for url in room_url_list:
            timestamp = random.randrange(10, 21)
            addInfo('本筆停留秒數: ' + str(timestamp) + ' 目前進度.... ' + str(count_rows) + '/ 約' + str(len(room_url_list) * checkDown) + '筆')
            time.sleep(timestamp)
            count_rows = count_rows + 1
            addr=''
            price=''
            size=''
            floor=''
            room_type=''
            now_environment=''
            car=''
            owner=''
            phone=''
            try:
                if comType.get() in WEBTYPE_A:
                    addr,price,size,floor,room_type,now_environment,car,owner,phone = getHouseData_A(url)
                if comType.get() in WEBTYPE_B:
                    addr,price,size,floor,room_type,now_environment,car,owner,phone = getHouseData_B(url)
                if comType.get() in WEBTYPE_C:
                    addr,price,size,floor,room_type,now_environment,car,owner,phone = getHouseData_C(url)
                if comType.get() in WEBTYPE_D:
                    addr,price,size,floor,room_type,now_environment,car,owner,phone = getHouseData_D(url)
            except:
                addInfo('running code:' + printLineFileFunc())

            # 準備Series 以及 append進DataFrame。值會放到相對印的column
            s = pd.Series([addr, price, size, floor, room_type, now_environment, car, owner, phone, phone],
                    index=["地址","價格","坪數","樓層","型態","現況","車位","屋主","電話","電話辨識"])
            # 因為 Series 沒有橫列的標籤, 所以加進去的時候一定要 ignore_index=True
            if len(s["屋主"]) <= 0:
                print('empty')
            else:
                df = df.append(s, ignore_index=True)
        print(df)

        sf = styleframe.StyleFrame(df)

        if i+1 < int(checkDown):
            browser.find_element_by_class_name('pageNext').send_keys(Keys.ESCAPE)
            browser.find_element_by_class_name('pageNext').click()
            time.sleep(1)

    sf.set_column_width_dict(col_width_dict={
        ("地址"): 65.5,
        ("價格","坪數","樓層","型態","現況","車位","屋主") : 20,
        ("電話", "電話辨識") : 25
     })

    all_rows = sf.row_indexes
    sf.set_row_height_dict(row_height_dict={
        all_rows[1:]: 30
    })

    output_file_name =  comType.get() + '_' +comCity.get() + '_'+  comArea.get() + now_today + '_' + str(time.time())+'.xlsx'

    sf.to_excel(output_file_name,
                sheet_name='Sheet1', #Create sheet
                right_to_left=False,
                columns_and_rows_to_freeze='A1',
                row_to_add_filters=0).save()

    addInfo('產出EXCEL中... ')

    if comType.get() in WEBTYPE_B:
        print()
    else:
        row = 0
        img_count = 0
        wb = ''
        wb = load_workbook(output_file_name)
        ws = wb.worksheets[0]

        img_path = "phone_img_"+ now_today
        dirFiles = os.listdir(img_path)

        # 第 8 欄為圖片
        for cell in list(ws.columns)[8]:
            if cell.value is None and img_count < len(dirFiles):
                # print('找到圖片 ', img_count, img_path + "/" + dirFiles[img_count])
                img = openpyxl.drawing.image.Image(img_path + "/" + dirFiles[img_count]) # create image instances

                # 圖片辨識
                im = Image.open(img_path + "/" + dirFiles[img_count])
                (x,y) = im.size #read image size
                x_s = 150 #define standard width
                y_s = y * x_s / x #calc height based on standard width
                out = im.resize((x_s, int (y_s)),Image.ANTIALIAS) #resize image with high-quality
                out.save(img_path + "/" + dirFiles[img_count])
                phone_t = pytesseract.image_to_string(out)

                # ws.add_image(img, 'H')
                c = str(row + 1)
                ws['J' + c] = ILLEGAL_CHARACTERS_RE.sub(r'', phone_t)
                ws.add_image(img, 'I' + c)
                img_count = img_count + 1
            row = row + 1

        wb.save(output_file_name)

    # 清除暫存檔案
    time.sleep(5)
    shutil.rmtree("phone_img_"+ now_today, ignore_errors=True)
    shutil.rmtree("debug.log", ignore_errors=True)
    browser.quit()
    addInfo('產生完成檔案:' + output_file_name)

def findTotal():
    global VERIFYPASS,varDownloadPage, df, browser, window

    btnAdd = tk.Button(window, text='匯出檔案', width=40, state=tk.DISABLED)
    btnAdd.place(x=150, y=120, width=100, height=20)

    if VERIFYPASS == False:
        addInfo('您尚未驗證,請先驗證')
        return

    if len(comCity.get()) == 0 or len(comArea.get()) == 0:
        addInfo('請選擇縣市或區域!!')
        return

    try:
        browser.quit()
    except:
        addInfo('running code:' + printLineFileFunc())

    df = pd.DataFrame(columns=["地址","價格","坪數","樓層","型態","現況","車位","屋主","電話","電話辨識"])

    addInfo('類型:' + comType.get() + ' 縣市:' + comCity.get() + ' 區域:' + comArea.get() + ' 搜尋...')
    window.after(500, setChrome)

def addInfo(sVar):
    dt1 = datetime.utcnow().replace(tzinfo=timezone.utc)
    now_time = dt1.astimezone(timezone(timedelta(hours=8)))
    dt_string = str(now_time.strftime("%Y/%m/%d %H:%M:%S"))
    # lstInfo.insert(0,'')
    result = dt_string + ' ' + sVar
    lstInfo.insert(0,result)
    window.update()

    # if os.path.isfile('runningcode.log'):
    #     f = open(r'runningcode.log')
    #     varVerify.set(result)
    #     f.close()
    # else:
    #     f = open("runningcode.log")
    #     f.write(result)
    #     f.close()

    print(result)

def printLineFileFunc():
    callerframerecord = inspect.stack()[1]
    frame = callerframerecord[0]
    info = inspect.getframeinfo(frame)
    filename = info.filename[info.filename.rfind('/')+1:]
    return str(info.lineno)

stdType = ('請先驗證')
comType = tt.Combobox(window, width=50, values=stdType)
comType.place(x=110, y=31, width=150, height=20)

def internet(event):
    cs = lstInfo.curselection()
    text=lstInfo.get(cs)
    if 'https' in text or 'http' in text:
        browser = webdriver.Chrome(executable_path=ChromeDriverManager().install(),chrome_options=options)
        browser.get(text.split(']')[1].strip())
    else:
        return

btnSetCityArea = tk.Button(window, text='確定縣市/區域', width=100, command=findTotal)
btnSetCityArea.place(x=270, y=60, width=100, height=20)

lstInfo = tk.Listbox(window, width=380)
lstInfo.place(x=10, y=140, width=380, height=180)
lstInfo.bind( "<Double-Button-1>" , internet)
addInfo('請先驗證後開始使用')

labFooter = tk.Label(window, text = '本篇文章僅作教學範例使用，實作本範例純屬個人行為\n作者不負任何法律責任，請點擊至知悉官網公告', justify=tk.LEFT, width=50, fg="blue", cursor="hand2")

labFooter.bind("<Button-1>", lambda e: webbrowser.open_new("https://help.591.com.tw/content/74/184/tw/%E6%9C%8D%E5%8B%99%E6%A2%9D%E6%AC%BE.html"))
labFooter.place(x=0, y=320, width=400, height=60)
window.mainloop()